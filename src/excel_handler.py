from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Tuple, Any
from pathlib import Path
from datetime import datetime
import os

MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_ROWS = 100000
MAX_PREVIEW_ROWS = 50


class ExcelHandler:
    @staticmethod
    def validate_excel_structure(file_path: str) -> Tuple[bool, str, List[Dict[str, Any]]]:
        try:
            if not os.path.exists(file_path):
                return False, f"文件不存在: {file_path}", []
            
            file_size = os.path.getsize(file_path)
            if file_size > MAX_FILE_SIZE:
                return False, f"文件大小超过限制（最大10MB，当前{(file_size / 1024 / 1024):.2f}MB）", []
            
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            if sheet.max_row < 2:
                wb.close()
                return False, "Excel文件至少需要包含表头和1行数据", []
            
            total_rows = sheet.max_row - 1
            if total_rows > MAX_ROWS:
                wb.close()
                return False, f"数据行数超过限制（最大10万行，当前{total_rows}行）", []
            
            if sheet.max_column < 2:
                wb.close()
                return False, "Excel文件至少需要2列（唯一标识列和待修改数据列）", []
            
            headers = [cell.value for cell in sheet[1]]
            if not headers[0] or not headers[1]:
                wb.close()
                return False, "前两列表头不能为空", []
            
            data_rows = []
            for row_idx in range(2, sheet.max_row + 1):
                row = [cell.value for cell in sheet[row_idx]]
                if row[0] is not None:
                    data_rows.append({
                        "row_num": row_idx,
                        "key_value": row[0],
                        "update_value": row[1] if len(row) > 1 else None
                    })
            wb.close()
            if not data_rows:
                return False, "Excel文件中没有有效数据行", []
            return True, f"验证通过，共 {len(data_rows)} 条数据", data_rows
        except Exception as e:
            return False, f"读取Excel文件失败: {str(e)}", []

    @staticmethod
    def check_duplicate_keys(data_rows: List[Dict[str, Any]]) -> Tuple[bool, List[Any]]:
        keys = [row["key_value"] for row in data_rows]
        seen = set()
        duplicates = []
        for key in keys:
            if key in seen and key not in duplicates:
                duplicates.append(key)
            seen.add(key)
        return len(duplicates) == 0, duplicates
    
    @staticmethod
    def get_key_values_from_excel(file_path: str) -> Tuple[bool, str, List[Any]]:
        """从Excel文件中获取所有唯一标识列的值"""
        try:
            if not os.path.exists(file_path):
                return False, f"文件不存在: {file_path}", []
            
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            
            if sheet.max_row < 2:
                wb.close()
                return False, "Excel文件至少需要包含表头和1行数据", []
            
            key_values = []
            for row_idx in range(2, sheet.max_row + 1):
                row = [cell.value for cell in sheet[row_idx]]
                if row[0] is not None:
                    key_values.append(row[0])
            
            wb.close()
            return True, f"获取了 {len(key_values)} 个唯一标识值", key_values
        except Exception as e:
            return False, f"读取Excel文件失败: {str(e)}", []
    
    @staticmethod
    def get_key_values_with_rows(file_path: str) -> Tuple[bool, str, List[Dict[str, Any]]]:
        """从Excel文件中获取所有唯一标识列的值及其行号"""
        try:
            if not os.path.exists(file_path):
                return False, f"文件不存在: {file_path}", []
            
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            
            if sheet.max_row < 2:
                wb.close()
                return False, "Excel文件至少需要包含表头和1行数据", []
            
            key_data = []
            for row_idx in range(2, sheet.max_row + 1):
                row = [cell.value for cell in sheet[row_idx]]
                if row[0] is not None:
                    key_data.append({
                        "row_num": row_idx,
                        "key_value": row[0]
                    })
            
            wb.close()
            return True, f"获取了 {len(key_data)} 条数据", key_data
        except Exception as e:
            return False, f"读取Excel文件失败: {str(e)}", []
    
    @staticmethod
    def validate_multi_column_structure(file_path: str, update_columns: List[str]) -> Tuple[bool, str, List[Dict[str, Any]]]:
        try:
            if not os.path.exists(file_path):
                return False, f"文件不存在: {file_path}", []
            
            file_size = os.path.getsize(file_path)
            if file_size > MAX_FILE_SIZE:
                return False, f"文件大小超过限制（最大10MB，当前{(file_size / 1024 / 1024):.2f}MB）", []
            
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            
            if sheet.max_row < 2:
                wb.close()
                return False, "Excel文件至少需要包含表头和1行数据", []
            
            total_rows = sheet.max_row - 1
            if total_rows > MAX_ROWS:
                wb.close()
                return False, f"数据行数超过限制（最大10万行，当前{total_rows}行）", []
            
            if sheet.max_column < 1 + len(update_columns):
                wb.close()
                return False, f"Excel文件至少需要{1 + len(update_columns)}列（1个唯一标识列 + {len(update_columns)}个待修改列）", []
            
            headers = [cell.value for cell in sheet[1]]
            key_column_name = headers[0]
            
            data_rows = []
            for row_idx in range(2, sheet.max_row + 1):
                row = [cell.value for cell in sheet[row_idx]]
                if row[0] is not None:
                    row_data = {
                        "row_num": row_idx,
                        "key_value": row[0]
                    }
                    for i, col_name in enumerate(update_columns):
                        row_data[col_name] = row[1 + i] if (1 + i) < len(row) else None
                    data_rows.append(row_data)
            
            wb.close()
            
            if not data_rows:
                return False, "Excel文件中没有有效数据行", []
            
            duplicate_valid, duplicates = ExcelHandler.check_duplicate_keys(data_rows)
            if not duplicate_valid:
                return False, f"Excel中唯一标识列存在重复值: {duplicates[:5]}", []
            
            return True, f"验证通过，共 {len(data_rows)} 条数据", data_rows
        except Exception as e:
            return False, f"读取Excel文件失败: {str(e)}", []
    
    @staticmethod
    def get_multi_column_preview(file_path: str, update_columns: List[str], max_rows: int = MAX_PREVIEW_ROWS) -> Tuple[bool, str, Dict[str, Any]]:
        try:
            file_size = os.path.getsize(file_path)
            if file_size > MAX_FILE_SIZE:
                return False, f"文件大小超过限制（最大10MB）", {}
            
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            
            headers = [str(cell.value) if cell.value else "" for cell in sheet[1]]
            preview_headers = [headers[0]] + update_columns[:min(len(update_columns), sheet.max_column - 1)]
            
            preview_rows = []
            actual_max = min(sheet.max_row, max_rows + 1)
            for row_idx in range(2, actual_max + 1):
                row = [cell.value for cell in sheet[row_idx]]
                preview_row = [row[0]] + [row[1 + i] if (1 + i) < len(row) else None for i in range(len(update_columns))]
                preview_rows.append(preview_row)
            
            wb.close()
            
            total_rows = sheet.max_row - 1
            
            return True, f"预览成功（显示前{MAX_PREVIEW_ROWS}行，共{total_rows}行）", {
                "headers": preview_headers,
                "rows": preview_rows[:max_rows],
                "total_rows": total_rows,
                "has_more": total_rows > MAX_PREVIEW_ROWS
            }
        except Exception as e:
            return False, f"预览失败: {str(e)}", {}

    @staticmethod
    def get_preview_data(file_path: str, max_rows: int = MAX_PREVIEW_ROWS) -> Tuple[bool, str, List[Dict[str, Any]]]:
        try:
            file_size = os.path.getsize(file_path)
            if file_size > MAX_FILE_SIZE:
                return False, f"文件大小超过限制（最大10MB）", {}
            
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            headers = [str(cell.value) if cell.value else "" for cell in sheet[1]]
            preview_rows = []
            actual_max = min(sheet.max_row, max_rows + 1)
            for row_idx in range(2, actual_max + 1):
                row = [cell.value for cell in sheet[row_idx]]
                preview_rows.append(row)
            wb.close()
            total_rows = sheet.max_row - 1 if hasattr(sheet, 'max_row') else len(preview_rows)
            return True, f"预览成功（显示前{MAX_PREVIEW_ROWS}行，共{total_rows}行）", {
                "headers": headers[:2],
                "rows": preview_rows[:max_rows],
                "total_rows": total_rows,
                "has_more": total_rows > MAX_PREVIEW_ROWS
            }
        except Exception as e:
            return False, f"预览失败: {str(e)}", {}

    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], file_path: str, sheet_name: str = "Sheet1") -> Tuple[bool, str]:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            if not data:
                return False, "没有数据可导出"
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            for col_idx in range(1, len(headers) + 1):
                max_length = max(
                    len(str(ws.cell(row=row, column=col_idx).value or ""))
                    for row in range(1, min(len(data) + 2, 1000))
                )
                ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = min(max_length + 2, 50)
            wb.save(file_path)
            return True, f"导出成功: {file_path}"
        except Exception as e:
            return False, f"导出失败: {str(e)}"

    @staticmethod
    def export_logs(log_entries: List[str], file_path: str) -> Tuple[bool, str]:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "更新日志"
            ws.append(["时间", "级别", "消息"])
            for entry in log_entries:
                entry = entry.strip()
                if entry:
                    try:
                        if entry.startswith("[") and "] " in entry:
                            time_part = entry[1:entry.index("]")]
                            level_msg = entry[entry.index("]") + 2:]
                            if " - " in level_msg:
                                level, message = level_msg.split(" - ", 1)
                            else:
                                level, message = "", level_msg
                            ws.append([time_part, level, message])
                        else:
                            ws.append(["", "", entry])
                    except Exception:
                        ws.append(["", "", entry])
            for col_idx in range(1, 4):
                ws.column_dimensions[chr(64 + col_idx)].width = 40
            wb.save(file_path)
            return True, f"日志导出成功: {file_path}"
        except Exception as e:
            return False, f"日志导出失败: {str(e)}"

    @staticmethod
    def export_failed_records(failed_records: List[Dict[str, Any]], file_path: str) -> Tuple[bool, str]:
        if not failed_records:
            return False, "没有失败记录可导出"
        return ExcelHandler.export_to_excel(failed_records, file_path, "失败记录")
