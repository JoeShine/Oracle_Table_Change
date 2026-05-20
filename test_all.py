#!/usr/bin/env python3
"""
Oracle数据批量修改工具 - 综合测试脚本
"""
import sys
import os
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

def test_config_manager():
    """测试配置管理模块"""
    print("\n[测试1] 配置管理模块...")
    from src.config_manager import ConfigManager

    config = ConfigManager()

    # 测试添加连接
    test_conn = {
        "name": "测试数据库",
        "host": "192.168.1.100",
        "port": 1521,
        "service": "TESTDB",
        "username": "TEST_USER",
        "password": "test123"
    }
    config.add_connection(test_conn)

    # 验证添加
    connections = config.get_connections()
    assert len(connections) == 1, "连接添加失败"
    assert connections[0]["name"] == "测试数据库"

    # 测试保存配置
    config.set_last_used(
        connection_name="测试数据库",
        target_table="TEST_TABLE",
        key_column="ID",
        update_column="VALUE"
    )

    # 验证保存
    last_used = config.get_last_used()
    assert last_used["target_table"] == "TEST_TABLE"

    print("✓ 配置管理测试通过")


def test_logger():
    """测试日志管理模块"""
    print("\n[测试2] 日志管理模块...")
    from src.logger import LogManager

    log = LogManager()

    # 测试日志记录
    log.info("测试信息")
    log.success("测试成功")
    log.error("测试错误")
    log.warning("测试警告")

    # 测试失败记录
    log.add_failed_record("KEY1", "VALUE1", "测试原因")
    failed = log.get_failed_records()
    assert len(failed) == 1
    assert failed[0]["key_value"] == "KEY1"

    # 测试获取日志
    logs = log.get_all_logs()
    assert len(logs) > 0

    print("✓ 日志管理测试通过")


def test_excel_handler():
    """测试Excel处理模块"""
    print("\n[测试3] Excel处理模块...")
    from src.excel_handler import ExcelHandler
    from openpyxl import Workbook
    import tempfile

    # 创建测试Excel文件
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'ID'
        ws['B1'] = 'VALUE'
        ws['A2'] = 1
        ws['B2'] = 'test1'
        ws['A3'] = 2
        ws['B3'] = 'test2'
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        # 测试验证
        success, msg, data = ExcelHandler.validate_excel_structure(tmp_path)
        assert success, f"验证失败: {msg}"
        assert len(data) == 2, "数据行数错误"

        # 测试重复检查
        dup_valid, dups = ExcelHandler.check_duplicate_keys(data)
        assert dup_valid, "重复检查失败"

        # 测试预览
        success, msg, preview = ExcelHandler.get_preview_data(tmp_path, 5)
        assert success
        assert len(preview["rows"]) <= 5

        # 测试导出失败记录
        failed_records = [
            {"key_value": 1, "update_value": "test", "reason": "错误", "timestamp": "2024-01-01"}
        ]
        export_path = tmp_path.replace('.xlsx', '_failed.xlsx')
        success, msg = ExcelHandler.export_failed_records(failed_records, export_path)
        assert success
        assert os.path.exists(export_path)
        os.unlink(export_path)

        print("✓ Excel处理测试通过")

    finally:
        os.unlink(tmp_path)


def test_db_connection():
    """测试数据库连接模块"""
    print("\n[测试4] 数据库连接模块...")
    from src.db_connection import DBConnection

    db = DBConnection()

    # 验证初始化
    assert db.connection is None
    assert not db.is_connected()

    print("✓ 数据库连接模块测试通过")


def test_data_updater():
    """测试数据更新器模块"""
    print("\n[测试5] 数据更新器模块...")
    from src.db_connection import DBConnection
    from src.logger import LogManager
    from src.data_updater import DataUpdater

    db = DBConnection()
    log = LogManager()
    updater = DataUpdater(db, log)

    # 验证初始化
    assert updater.db is not None
    assert updater.log is not None
    assert updater.temp_table_name is None

    print("✓ 数据更新器模块测试通过")


def test_gui_import():
    """测试GUI模块导入"""
    print("\n[测试6] GUI模块导入...")
    try:
        from src.gui import OracleBatchUpdaterGUI
        print("✓ GUI模块导入测试通过")
    except ImportError as e:
        if 'DISPLAY' in str(e) or 'no display' in str(e).lower():
            print("✓ GUI模块导入测试通过（无图形环境）")
        else:
            raise


def main():
    """主测试函数"""
    print("=" * 60)
    print("Oracle数据批量修改工具 - 综合测试")
    print("=" * 60)

    try:
        test_config_manager()
        test_logger()
        test_excel_handler()
        test_db_connection()
        test_data_updater()
        test_gui_import()

        print("\n" + "=" * 60)
        print("✓ 所有测试通过！")
        print("=" * 60)
        return 0

    except Exception as e:
        print("\n" + "=" * 60)
        print(f"✗ 测试失败: {str(e)}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
