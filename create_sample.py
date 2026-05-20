from openpyxl import Workbook
from datetime import datetime

def create_sample_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "用户邮箱更新"

    ws['A1'] = 'USER_ID'
    ws['B1'] = 'EMAIL'

    ws['A2'] = 1001
    ws['B2'] = 'zhangsan@example.com'

    ws['A3'] = 1002
    ws['B3'] = 'lisi@example.com'

    ws['A4'] = 1003
    ws['B4'] = 'wangwu@example.com'

    ws['A5'] = 1004
    ws['B5'] = 'zhaoliu@example.com'

    ws['A6'] = 1005
    ws['B6'] = 'sunqi@example.com'

    filename = f"sample_update_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"示例Excel文件已创建: {filename}")
    return filename

if __name__ == "__main__":
    create_sample_excel()
