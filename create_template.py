#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Oracle数据批量修改工具 - Excel导入模板生成器
生成美观的Excel模板，带首行配色和首行冻结
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "数据更新模板"
    
    # 定义表头
    headers = ["唯一标识列", "更新列1", "更新列2", "更新列3", "更新列4", "备注"]
    
    # 首行配色和样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    # 添加示例数据
    sample_data = [
        ["001", "张三", "28", "技术部", "工程师", "正常员工"],
        ["002", "李四", "32", "市场部", "经理", "正常员工"],
        ["003", "王五", "25", "财务部", "会计", "正常员工"],
        ["004", "赵六", "30", "人事部", "专员", "正常员工"],
        ["005", "钱七", "35", "技术部", "架构师", "核心员工"],
    ]
    
    # 设置示例数据样式
    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            # 隔行变色
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # 首行冻结
    ws.freeze_panes = "A2"
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    for i in range(2, len(sample_data) + 2):
        ws.row_dimensions[i].height = 20
    
    # 添加说明页
    ws_guide = wb.create_sheet("使用说明", 0)
    
    guide_header = ["项目", "内容"]
    guide_header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    guide_header_fill = PatternFill(start_color="3592C4", end_color="3592C4", fill_type="solid")
    
    ws_guide.cell(row=1, column=1, value="Oracle数据批量修改工具 - Excel模板使用说明")
    ws_guide.merge_cells("A1:B1")
    guide_title = ws_guide.cell(row=1, column=1)
    guide_title.font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
    guide_title.fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")
    guide_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[1].height = 30
    
    guide_content = [
        ["模板说明", "本模板用于Oracle数据批量修改工具的数据导入"],
        ["必填字段", "第1列必须为唯一标识列，不能为空"],
        ["更新字段", "第2列及以后为待更新的字段，根据实际需要填写"],
        ["文件格式", "建议使用.xlsx格式（Excel 2007及以上）"],
        ["文件大小", "建议不超过10MB，数据不超过10万行"],
        ["数据类型", "请确保数据类型与数据库表字段类型一致"],
        ["特殊字符", "避免使用特殊字符或中文符号"],
        ["备份数据", "更新前系统会自动备份，请确认后再操作"],
    ]
    
    for row_idx, content in enumerate(guide_content, 2):
        cell1 = ws_guide.cell(row=row_idx, column=1, value=content[0])
        cell2 = ws_guide.cell(row=row_idx, column=2, value=content[1])
        
        for cell in [cell1, cell2]:
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        ws_guide.column_dimensions["A"].width = 20
        ws_guide.column_dimensions["B"].width = 50
        ws_guide.row_dimensions[row_idx].height = 20
    
    # 保存文件
    output_file = "Oracle数据批量修改工具_导入模板.xlsx"
    wb.save(output_file)
    print(f"Excel导入模板已生成: {output_file}")
    return output_file

if __name__ == "__main__":
    create_excel_template()
