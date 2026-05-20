#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""创建测试用Excel文件 - 用于一致性校验测试"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "测试数据"

# 添加表头
headers = ["ID", "姓名", "年龄", "部门"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 添加示例数据 - 包含数据库中不存在的ID
data = [
    [1001, "张三", 28, "技术部"],
    [1002, "李四", 30, "市场部"],
    [1003, "王五", 32, "财务部"],
    [9999, "不存在的人", 30, "未知部门"],  # 这个ID在数据库中不存在
    [1004, "赵六", 29, "人事部"],
    [1005, "钱七", 31, "技术部"],
    [8888, "测试人员", 25, "测试部"],  # 这个ID在数据库中不存在
    [1006, "孙八", 27, "市场部"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 设置列宽
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 15

# 保存文件
wb.save("test_consistency_data.xlsx")
print("✅ 测试文件已创建: test_consistency_data.xlsx")
print("   包含数据库中不存在的ID测试数据")
print("\n数据预览:")
for row in data:
    print(f"   {row}")
